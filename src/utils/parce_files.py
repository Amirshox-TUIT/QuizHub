import os

from docx import Document
import pdfplumber
from openpyxl import load_workbook


def parse_docx(file) -> tuple[str, list[tuple[str, list[str]]]]:
    """
    Возвращает: title_quiz, list of (question_text, [answers])
    """
    doc = Document(file)
    rows = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    if not rows:
        raise ValueError("Документ пустой")

    quiz_title = rows[0]
    questions = []

    idx = 1
    while idx < len(rows):
        question_text = rows[idx]
        idx += 1
        answers_texts = []
        while idx < len(rows) and len(answers_texts) < 4:
            answers_texts.append(rows[idx])
            idx += 1
        if len(answers_texts) == 4:
            questions.append((question_text, answers_texts))
    os.remove(file)
    return quiz_title, questions


def parse_pdf(file) -> tuple[str, list[tuple[str, list[str]]]]:
    text_lines = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text_lines.extend([line.strip() for line in page.extract_text().split("\n") if line.strip()])

    if not text_lines:
        raise ValueError("PDF пустой")

    quiz_title = text_lines[0]
    questions = []

    idx = 1
    while idx < len(text_lines):
        question_text = text_lines[idx]
        idx += 1
        answers_texts = []
        while idx < len(text_lines) and len(answers_texts) < 4:
            answers_texts.append(text_lines[idx])
            idx += 1
        if len(answers_texts) == 4:
            questions.append((question_text, answers_texts))

    os.remove(file)
    return quiz_title, questions


def parse_xlsx(file) -> tuple[str, list[tuple[str, list[str]]]]:
    wb = load_workbook(file)
    sheet = wb.active
    rows = [row[0].value.strip() for row in sheet.iter_rows(min_row=1) if row[0].value and row[0].value.strip()]

    if not rows:
        raise ValueError("Файл пустой")

    quiz_title = rows[0]
    questions = []

    idx = 1
    while idx < len(rows):
        question_text = rows[idx]
        idx += 1
        answers_texts = []
        while idx < len(rows) and len(answers_texts) < 4:
            answers_texts.append(rows[idx])
            idx += 1
        if len(answers_texts) == 4:
            questions.append((question_text, answers_texts))

    os.remove(file)
    return quiz_title, questions


